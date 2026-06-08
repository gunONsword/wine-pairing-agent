from __future__ import annotations

import argparse
import json
import os
import shutil
import sys
from datetime import datetime
from pathlib import Path
from typing import Any

from openpyxl import load_workbook
from pydantic import BaseModel, Field


PROJECT_ROOT = Path(__file__).resolve().parents[1]
if str(PROJECT_ROOT) not in sys.path:
    sys.path.insert(0, str(PROJECT_ROOT))

from app.config import load_dotenv  # noqa: E402
from services.llm_service import build_llm, invoke_json_model  # noqa: E402


DEFAULT_WORKBOOK = PROJECT_ROOT / "data" / "202605_ペアリング評価用シート(第2回).xlsx"
DEFAULT_RESULT_COLUMN = "agent_pairing_evaluation"


class PairingEvaluation(BaseModel):
    score: int = Field(ge=1, le=5)
    reason: str
    positive_points: list[str] = Field(default_factory=list)
    concerns: list[str] = Field(default_factory=list)


def parse_args() -> argparse.Namespace:
    parser = argparse.ArgumentParser(
        description="Evaluate dish/wine pairings in 評価シート and write results back row by row.",
    )
    parser.add_argument(
        "--workbook",
        type=Path,
        default=DEFAULT_WORKBOOK,
        help=f"Source Excel workbook path. Default: {DEFAULT_WORKBOOK}",
    )
    parser.add_argument(
        "--output-workbook",
        type=Path,
        default=None,
        help="Copied workbook path to write results into. Defaults to a timestamped copy next to the source.",
    )
    parser.add_argument("--eval-sheet", default="評価シート")
    parser.add_argument("--dish-sheet", default="料理")
    parser.add_argument("--wine-sheet", default="ワイン")
    parser.add_argument(
        "--result-column",
        default=DEFAULT_RESULT_COLUMN,
        help="Column name to create/use for evaluation output.",
    )
    parser.add_argument(
        "--env-file",
        type=Path,
        default=None,
        help="Env file for MINIMAX_API_KEY. Defaults to .env, or env if .env is missing.",
    )
    parser.add_argument("--limit", type=int, default=None, help="Evaluate at most N pending rows.")
    parser.add_argument("--overwrite", action="store_true", help="Re-evaluate rows that already have results.")
    return parser.parse_args()


def normalize_key(value: Any) -> str:
    if value is None:
        return ""
    if isinstance(value, float) and value.is_integer():
        return str(int(value))
    return str(value).strip()


def normalized_header(value: Any) -> str:
    return str(value or "").strip()


def header_map(sheet) -> dict[str, int]:
    return {normalized_header(cell.value): cell.column for cell in sheet[1] if normalized_header(cell.value)}


def find_header(headers: dict[str, int], *candidates: str, startswith: str | None = None) -> int:
    for candidate in candidates:
        if candidate in headers:
            return headers[candidate]
    if startswith:
        for header, column in headers.items():
            if header.startswith(startswith):
                return column
    raise KeyError(f"Missing required column: {candidates or startswith}")


def ensure_result_column(sheet, column_name: str) -> int:
    headers = header_map(sheet)
    if column_name in headers:
        return headers[column_name]

    column = sheet.max_column + 1
    sheet.cell(row=1, column=column, value=column_name)
    return column


def row_dict(sheet, row: int) -> dict[str, Any]:
    result: dict[str, Any] = {}
    for cell in sheet[1]:
        key = normalized_header(cell.value)
        if key:
            result[key] = sheet.cell(row=row, column=cell.column).value
    return result


def build_dish_index(sheet) -> dict[str, dict[str, Any]]:
    headers = header_map(sheet)
    id_col = find_header(headers, startswith="料理ID")
    dishes: dict[str, dict[str, Any]] = {}

    for row in range(2, sheet.max_row + 1):
        dish_id = normalize_key(sheet.cell(row=row, column=id_col).value)
        if dish_id:
            dishes[dish_id] = row_dict(sheet, row)

    return dishes


def build_wine_index(sheet) -> dict[str, dict[str, Any]]:
    headers = header_map(sheet)
    id_col = find_header(headers, startswith="商品コード")
    wines: dict[str, dict[str, Any]] = {}

    for row in range(2, sheet.max_row + 1):
        item_id = normalize_key(sheet.cell(row=row, column=id_col).value)
        if not item_id:
            continue
        wine = row_dict(sheet, row)
        wines[item_id] = wine
        wines.setdefault(item_id.lstrip("0") or "0", wine)

    return wines


def compact_dict(data: dict[str, Any], max_value_length: int = 1200) -> dict[str, str]:
    compacted: dict[str, str] = {}
    for key, value in data.items():
        if value is None or value == "":
            continue
        text = str(value).strip()
        if len(text) > max_value_length:
            text = text[:max_value_length] + "...[truncated]"
        compacted[key] = text
    return compacted


def build_prompt(dish: dict[str, Any], wine: dict[str, Any]) -> str:
    dish_json = json.dumps(compact_dict(dish), ensure_ascii=False, indent=2)
    wine_json = json.dumps(compact_dict(wine), ensure_ascii=False, indent=2)
    return f"""
あなたはワインと料理のペアリングを評価するソムリエAIです。
次の料理とワインの相性を、1から5の整数スコアで評価してください。

評価基準:
- 1: 料理とワインの相性が悪い。味、香り、重さ、酸味、渋みなどがぶつかる。
- 3: 可もなく不可もない。成立はするが、強い必然性はない。
- 5: とても相性が良い。料理の特徴をワインが引き立て、ペアリングとして説得力がある。

料理:
{dish_json}

ワイン:
{wine_json}

必ず次の JSON オブジェクトだけを返してください。
{{
  "score": 1,
  "reason": "評価理由を日本語で簡潔に書く",
  "positive_points": ["良い点1", "良い点2"],
  "concerns": ["懸念点1"]
}}
""".strip()


def format_result(evaluation: PairingEvaluation) -> str:
    lines = [
        f"score: {evaluation.score}/5",
        f"reason: {evaluation.reason}",
    ]
    if evaluation.positive_points:
        lines.append("positive_points: " + " / ".join(evaluation.positive_points))
    if evaluation.concerns:
        lines.append("concerns: " + " / ".join(evaluation.concerns))
    return "\n".join(lines)


def choose_env_file(value: Path | None) -> Path | None:
    if value:
        return value
    dot_env = PROJECT_ROOT / ".env"
    plain_env = PROJECT_ROOT / "env"
    if dot_env.exists():
        return dot_env
    if plain_env.exists():
        return plain_env
    return None


def create_working_copy(source: Path, output: Path | None) -> Path:
    if not source.exists():
        raise FileNotFoundError(f"Workbook not found: {source}")

    if output is None:
        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        output = source.with_name(f"{source.stem}_evaluated_{timestamp}{source.suffix}")
    else:
        output = output.resolve()

    if output.exists():
        raise FileExistsError(f"Output workbook already exists: {output}")

    output.parent.mkdir(parents=True, exist_ok=True)
    shutil.copy2(source, output)
    return output


def main() -> None:
    args = parse_args()
    source_workbook_path = args.workbook.resolve()
    workbook_path = create_working_copy(source_workbook_path, args.output_workbook)
    print(f"copied workbook: {source_workbook_path} -> {workbook_path}")

    env_file = choose_env_file(args.env_file)
    if env_file:
        load_dotenv(env_file)

    if not os.getenv("MINIMAX_API_KEY"):
        raise RuntimeError("MINIMAX_API_KEY is required. Put it in .env/env or pass --env-file.")

    workbook = load_workbook(workbook_path)
    eval_sheet = workbook[args.eval_sheet]
    dish_sheet = workbook[args.dish_sheet]
    wine_sheet = workbook[args.wine_sheet]

    eval_headers = header_map(eval_sheet)
    dish_id_col = find_header(eval_headers, "dish_id")
    item_id_col = find_header(eval_headers, "item_id", "item_cd")
    result_col = ensure_result_column(eval_sheet, args.result_column)

    dishes = build_dish_index(dish_sheet)
    wines = build_wine_index(wine_sheet)
    llm = build_llm()

    evaluated = 0
    for row in range(2, eval_sheet.max_row + 1):
        if args.limit is not None and evaluated >= args.limit:
            break

        result_cell = eval_sheet.cell(row=row, column=result_col)
        if result_cell.value and not args.overwrite:
            continue

        dish_id = normalize_key(eval_sheet.cell(row=row, column=dish_id_col).value)
        item_id = normalize_key(eval_sheet.cell(row=row, column=item_id_col).value)

        dish = dishes.get(dish_id)
        wine = wines.get(item_id) or wines.get(item_id.lstrip("0") or "0")

        if not dish or not wine:
            missing = []
            if not dish:
                missing.append(f"dish_id={dish_id}")
            if not wine:
                missing.append(f"item_id={item_id}")
            result_cell.value = "ERROR: missing " + ", ".join(missing)
            workbook.save(workbook_path)
            print(f"row {row}: missing data, saved error")
            continue

        prompt = build_prompt(dish, wine)
        evaluation = invoke_json_model(llm, prompt, PairingEvaluation)
        result_cell.value = format_result(evaluation)
        workbook.save(workbook_path)
        evaluated += 1
        print(f"row {row}: score={evaluation.score}/5 saved")

    print(f"done: evaluated {evaluated} row(s)")


if __name__ == "__main__":
    main()
